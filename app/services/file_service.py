from datetime import datetime
from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.utils.constants import REQUIRED_COLUMNS


def save_uploaded_file(uploaded_file: FileStorage, upload_folder: Path) -> Path:
    """
    Сохраняет загруженный пользователем файл в папку storage/uploads.
    Чтобы имена не конфликтовали, к имени добавляется дата и случайный идентификатор.
    """

    upload_folder.mkdir(parents=True, exist_ok=True)

    original_filename = secure_filename(uploaded_file.filename)
    extension = Path(original_filename).suffix.lower()

    unique_name = (
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_"
        f"{uuid4().hex[:8]}"
        f"{extension}"
    )

    saved_path = upload_folder / unique_name
    uploaded_file.save(saved_path)

    return saved_path


def read_data_file(file_path: Path) -> pd.DataFrame:
    """
    Читает Excel или CSV-файл и возвращает pandas DataFrame.
    """

    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return pd.read_excel(file_path, engine="openpyxl")

    if suffix == ".xls":
        raise ValueError(
            "Формат .xls пока не поддерживается текущими зависимостями проекта. "
            "Сохраните файл в формате .xlsx или .csv."
        )

    if suffix == ".csv":
        return _read_csv_file(file_path)

    raise ValueError("Неизвестный формат файла.")


def _read_csv_file(file_path: Path) -> pd.DataFrame:
    """
    Читает CSV-файл. Пробует несколько популярных кодировок.
    """

    encodings = ["utf-8-sig", "utf-8", "cp1251"]

    last_error = None

    for encoding in encodings:
        try:
            return pd.read_csv(
                file_path,
                encoding=encoding,
                sep=None,
                engine="python"
            )
        except UnicodeDecodeError as error:
            last_error = error

    raise ValueError(
        "Не удалось прочитать CSV-файл. "
        "Попробуйте сохранить файл в кодировке UTF-8."
    ) from last_error


def create_template_file(template_path: Path) -> Path:
    """
    Создаёт Excel-шаблон для загрузки данных.
    """

    template_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Данные студентов"

    worksheet.append(REQUIRED_COLUMNS)

    sample_rows = [
        ["Иванов И.И.", "ИС-21", "Математика", "Очное обучение", 86],
        ["Петров П.П.", "ИС-21", "Математика", "Онлайн-курс", 74],
        ["Сидорова А.А.", "ИС-22", "Математика", "Смешанное обучение", 91],
    ]

    for row in sample_rows:
        worksheet.append(row)

    header_fill = PatternFill(
        start_color="DCE9FB",
        end_color="DCE9FB",
        fill_type="solid"
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    column_widths = {
        "A": 24,
        "B": 14,
        "C": 22,
        "D": 24,
        "E": 22,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    workbook.save(template_path)

    return template_path